"""
MFS Scam Annotation — Dataset Preparation Script
=================================================
Run this from your mfs_scam_research project folder:

    python annotation_package/prepare_annotation_set.py

Outputs (written to ./annotation_package/):
    annotation_batch.csv          — 600 messages ready for annotation
    annotation_sheet_A1.xlsx      — spreadsheet for Annotator 1
    annotation_sheet_A2.xlsx      — spreadsheet for Annotator 2
    kappa_calculator.py           — run after both annotators finish
"""

import csv
import json
import random
import os
import sys
from pathlib import Path
from collections import Counter

random.seed(42)  # reproducible

BASE   = Path(__file__).parent.parent   # mfs_scam_research/
OUT    = Path(__file__).parent          # annotation_package/
OUT.mkdir(exist_ok=True)

TARGET = 600   # total messages for annotation

# ── 7 annotation categories ──────────────────────────────────────────────────

CATEGORIES = [
    "S1_AccountCompromise",   # OTP theft, hacking, unauthorised access
    "S2_FakePromo",           # Fake cashback, fake offers, fake bonus
    "S3_LotteryPrize",        # Lottery won, prize claim scams
    "S4_LoanJob",             # Fake loan, fake job offer
    "S5_Impersonation",       # Fake agent, fake customer care
    "S6_PhishingLink",        # Phishing URL, fake app, clone website
    "L0_Ham",                 # Legitimate / not a scam
]

# ── Source 1: Mendeley financial scams dataset ───────────────────────────────

def load_financial_scams():
    path = BASE / "stream_a_datasets" / "clean_dataset_1.csv"
    if not path.exists():
        print(f"  [!] Not found: {path}")
        return []
    rows = []
    with open(path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            text = (r.get("message") or r.get("text") or "").strip()
            label = (r.get("label") or "").strip().lower()
            if len(text) > 15:
                rows.append({
                    "source":   "mendeley_financial",
                    "text":     text,
                    "orig_label": label,
                    "prelabel": "S1_AccountCompromise" if label == "scam" else "L0_Ham",
                })
    print(f"  Loaded {len(rows)} from Mendeley financial scams dataset")
    return rows

# ── Source 2: BanglaBarta dataset ───────────────────────────────────────────

def load_banglabarta():
    path = BASE / "stream_a_datasets" / "clean_dataset_2.csv"
    if not path.exists():
        print(f"  [!] Not found: {path}")
        return []
    rows = []
    prelabel_map = {
        "smish":  "S6_PhishingLink",
        "promo":  "S2_FakePromo",
        "normal": "L0_Ham",
    }
    with open(path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            text = (r.get("text") or "").strip()
            label = (r.get("label") or "").strip().lower()
            if len(text) > 15:
                rows.append({
                    "source":   "banglabarta",
                    "text":     text,
                    "orig_label": label,
                    "prelabel": prelabel_map.get(label, "Unclassified"),
                })
    print(f"  Loaded {len(rows)} from BanglaBarta dataset")
    return rows

# ── Source 3: App review fraud messages ─────────────────────────────────────

def load_app_reviews():
    rows = []
    fraud_dir = BASE / "app_reviews"
    # try both possible locations
    for candidate in [fraud_dir, BASE / "stream_e_app_reviews"]:
        for json_file in candidate.glob("*fraud*.json"):
            try:
                with open(json_file, encoding="utf-8") as f:
                    data = json.load(f)
                for r in data:
                    text = (r.get("text") or r.get("content") or "").strip()
                    if len(text) > 15:
                        rows.append({
                            "source":   f"app_review_{r.get('app','')}",
                            "text":     text,
                            "orig_label": "scam",
                            "prelabel": r.get("prelabel", "S1_AccountCompromise"),
                        })
            except Exception as e:
                print(f"  [!] Could not load {json_file}: {e}")
    # also try CSV
    for candidate in [BASE / "app_reviews", BASE]:
        for csv_file in candidate.glob("*reviews*.csv"):
            try:
                with open(csv_file, encoding="utf-8") as f:
                    for r in csv.DictReader(f):
                        if str(r.get("fraud_flag","")).lower() in ("true","1","yes"):
                            text = (r.get("text") or "").strip()
                            if len(text) > 15:
                                rows.append({
                                    "source":   f"app_review_{r.get('app','')}",
                                    "text":     text,
                                    "orig_label": "scam",
                                    "prelabel": r.get("prelabel") or "S1_AccountCompromise",
                                })
            except Exception as e:
                print(f"  [!] Could not load {csv_file}: {e}")
    print(f"  Loaded {len(rows)} from app reviews (fraud-flagged)")
    return rows

# ── Source 4: Telegram fraud messages ───────────────────────────────────────

def load_telegram():
    rows = []
    for candidate in [BASE / "telegram_data", BASE / "stream_d_social_media" / "telegram_data"]:
        for sub in ["fraud", "processed", "csv", "raw"]:
            d = candidate / sub
            if not d.exists():
                continue
            for jf in d.glob("*fraud*.json"):
                try:
                    with open(jf, encoding="utf-8") as f:
                        data = json.load(f)
                    for r in data:
                        text = (r.get("text") or "").strip()
                        if len(text) > 15:
                            rows.append({
                                "source":   f"telegram_{r.get('channel','')}",
                                "text":     text,
                                "orig_label": "scam",
                                "prelabel": r.get("prelabel", "S5_Impersonation"),
                            })
                except Exception as e:
                    print(f"  [!] {jf}: {e}")
            for cf in d.glob("*.csv"):
                try:
                    with open(cf, encoding="utf-8") as f:
                        for r in csv.DictReader(f):
                            if str(r.get("fraud_flag","")).lower() in ("true","1","yes"):
                                text = (r.get("text") or "").strip()
                                if len(text) > 15:
                                    rows.append({
                                        "source":   f"telegram_{r.get('channel','')}",
                                        "text":     text,
                                        "orig_label": "scam",
                                        "prelabel": r.get("prelabel","S5_Impersonation"),
                                    })
                except Exception as e:
                    print(f"  [!] {cf}: {e}")
    print(f"  Loaded {len(rows)} from Telegram (fraud-flagged)")
    return rows

# ── Sampling strategy ────────────────────────────────────────────────────────

def stratified_sample(pool, n):
    """Sample n items, keeping source diversity. Dedup by text first."""
    seen = set()
    unique = []
    for r in pool:
        key = r["text"][:80].lower()
        if key not in seen:
            seen.add(key)
            unique.append(r)

    # Group by source
    by_source = {}
    for r in unique:
        src = r["source"].split("_")[0]  # coarse grouping
        by_source.setdefault(src, []).append(r)

    # Proportional allocation
    total = len(unique)
    result = []
    for src, items in by_source.items():
        quota = max(1, round(len(items) / total * n))
        result.extend(random.sample(items, min(quota, len(items))))

    # Top up or trim to exactly n
    remaining = [r for r in unique if r not in result]
    random.shuffle(remaining)
    result.extend(remaining)
    result = result[:n]
    random.shuffle(result)
    return result

# ── Write annotation CSV ─────────────────────────────────────────────────────

def write_batch_csv(messages):
    out = OUT / "annotation_batch.csv"
    fields = ["msg_id","source","text","prelabel","annotator1_label","annotator2_label","notes"]
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for i, m in enumerate(messages, 1):
            w.writerow({
                "msg_id": f"MSG{i:04d}",
                "source": m["source"],
                "text": m["text"],
                "prelabel": m["prelabel"],
                "annotator1_label": "",
                "annotator2_label": "",
                "notes": "",
            })
    print(f"\n  Saved {len(messages)} messages → {out.name}")
    return out

# ── Write annotator Excel sheets ─────────────────────────────────────────────

def write_excel_sheets(messages):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("\n  [!] openpyxl not installed. Run: pip install openpyxl --break-system-packages")
        print("      Skipping Excel output. annotation_batch.csv was still saved.")
        return

    PURPLE = "461D7C"
    GOLD   = "FDD023"
    LIGHT  = "F3EFF8"
    GREEN  = "E8F5E9"

    cat_list = ",".join(CATEGORIES)
    validation = DataValidation(
        type="list",
        formula1=f'"{cat_list}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid category",
        error=f"Choose one of: {', '.join(CATEGORIES)}"
    )

    for annotator_num in [1, 2]:
        wb = openpyxl.Workbook()

        # ── Instructions sheet ──────────────────────────────────────────────
        ws_inst = wb.active
        ws_inst.title = "INSTRUCTIONS"

        inst_rows = [
            ("MFS SCAM ANNOTATION TASK", True, 16, PURPLE),
            ("", False, 11, "000000"),
            ("ANNOTATOR INFORMATION", True, 12, PURPLE),
            (f"Annotator Number: {annotator_num}", False, 11, "000000"),
            ("Your Name: ___________________________", False, 11, "000000"),
            ("Date Started: ___________________________", False, 11, "000000"),
            ("Date Completed: ___________________________", False, 11, "000000"),
            ("", False, 11, "000000"),
            ("TASK DESCRIPTION", True, 12, PURPLE),
            ("You will label 600 text messages into one of 7 categories.", False, 11, "000000"),
            ("Each message is a real MFS-related text (from app reviews, SMS datasets, or Telegram).", False, 11, "000000"),
            ("Go to the ANNOTATION sheet. For each message, select a label from the dropdown in column E.", False, 11, "000000"),
            ("If you are unsure, add a note in column F.", False, 11, "000000"),
            ("Do NOT look at the 'prelabel' column (D) until after you assign your label.", False, 11, "000000"),
            ("", False, 11, "000000"),
            ("THE 7 CATEGORIES", True, 12, PURPLE),
            ("S1_AccountCompromise — OTP theft, account hacking, unauthorised transactions, PIN/password stolen", False, 11, "000000"),
            ("  Example: 'Someone transferred Tk 5000 from my bKash without my permission. My OTP was stolen.'", False, 10, "555555"),
            ("S2_FakePromo — Fake cashback offers, fake bonus, fake promotional messages", False, 11, "000000"),
            ("  Example: 'You have won Tk 500 cashback! Send your bKash number to claim.'", False, 10, "555555"),
            ("S3_LotteryPrize — Lottery scams, prize claim fraud, contest fraud", False, 11, "000000"),
            ("  Example: 'Congratulations! You won Tk 10,000 in the Nagad Lucky Draw. Call now to claim.'", False, 10, "555555"),
            ("S4_LoanJob — Fake loan offers, fake job offers, advance fee fraud", False, 11, "000000"),
            ("  Example: 'Easy loan Tk 50,000 via bKash. Pay Tk 500 processing fee first.'", False, 10, "555555"),
            ("S5_Impersonation — Fake bKash/Nagad agent, fake customer care, fake employee", False, 11, "000000"),
            ("  Example: 'I am calling from Nagad head office. Please share your PIN to verify your account.'", False, 10, "555555"),
            ("S6_PhishingLink — Fake website link, fake app, clone site, malware link", False, 11, "000000"),
            ("  Example: 'Your bKash account will be suspended. Verify now: http://bkash-verify.xyz'", False, 10, "555555"),
            ("L0_Ham — Legitimate message. Not a scam. Normal transaction, genuine complaint, or unrelated.", False, 11, "000000"),
            ("  Example: 'bKash is easy to use. I send money every month. 5 stars.'", False, 10, "555555"),
            ("", False, 11, "000000"),
            ("LABELLING RULES", True, 12, PURPLE),
            ("1. Assign ONLY ONE label per message.", False, 11, "000000"),
            ("2. Choose the PRIMARY scam type if a message fits multiple categories.", False, 11, "000000"),
            ("3. If a message is ambiguous, use your best judgment and note it in column F.", False, 11, "000000"),
            ("4. Label L0_Ham only if you are confident the message is NOT a scam.", False, 11, "000000"),
            ("5. Work independently — do not discuss labels with the other annotator until both are done.", False, 11, "000000"),
            ("", False, 11, "000000"),
            ("TARGET: Complete all 600 messages. Estimated time: 3–5 hours (work in sessions).", True, 11, PURPLE),
            ("CONTACT: saidur@lsu.edu for any questions.", False, 11, "000000"),
        ]

        for row_idx, (text, bold, size, color) in enumerate(inst_rows, 1):
            cell = ws_inst.cell(row=row_idx, column=1, value=text)
            cell.font = Font(bold=bold, size=size, color=color,
                             name="Calibri" if not bold else "Calibri")
            cell.alignment = Alignment(wrap_text=True)

        ws_inst.column_dimensions["A"].width = 110

        # ── Annotation sheet ────────────────────────────────────────────────
        ws = wb.create_sheet("ANNOTATION")
        ws.add_data_validation(validation)

        headers = ["MSG_ID", "SOURCE", "TEXT", "PRE-LABEL (guide only)", "YOUR LABEL ▼", "NOTES (optional)"]
        col_widths = [10, 18, 80, 26, 26, 30]

        # Header row
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=PURPLE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 30

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Category colours for pre-label column
        cat_colors = {
            "S1_AccountCompromise": "FFCDD2",
            "S2_FakePromo":         "FFF9C4",
            "S3_LotteryPrize":      "F3E5F5",
            "S4_LoanJob":           "E3F2FD",
            "S5_Impersonation":     "FCE4D6",
            "S6_PhishingLink":      "FFE0B2",
            "L0_Ham":               "E8F5E9",
            "Unclassified":         "F5F5F5",
        }

        for row_idx, m in enumerate(messages, 2):
            row_fill = PatternFill("solid", fgColor=LIGHT if row_idx % 2 == 0 else "FFFFFF")
            vals = [
                f"MSG{(row_idx-1):04d}",
                m["source"],
                m["text"],
                m["prelabel"],
                "",
                "",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 4:  # prelabel — colour coded
                    color = cat_colors.get(m["prelabel"], "F5F5F5")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(size=9, color="555555", name="Calibri", italic=True)
                elif col == 5:  # label dropdown
                    cell.fill = PatternFill("solid", fgColor="FFFDE7")
                    cell.font = Font(size=11, bold=True, name="Calibri")
                    validation.sqref = validation.sqref  # will add below
                else:
                    cell.fill = row_fill
                    cell.font = Font(size=10, name="Calibri")

            # Apply dropdown to column E
            ws.cell(row=row_idx, column=5).data_type = "s"

        # Add validation to entire E column (rows 2 onwards)
        validation.sqref = f"E2:E{len(messages)+1}"

        ws.freeze_panes = "A2"
        ws.sheet_view.zoomScale = 90

        # ── Summary sheet ───────────────────────────────────────────────────
        ws_sum = wb.create_sheet("SUMMARY")
        ws_sum["A1"] = "Category"
        ws_sum["B1"] = "Count"
        ws_sum["C1"] = "Formula (auto-updates)"
        ws_sum["A1"].font = Font(bold=True, color="FFFFFF", name="Calibri")
        ws_sum["B1"].font = Font(bold=True, color="FFFFFF", name="Calibri")
        ws_sum["C1"].font = Font(bold=True, color="FFFFFF", name="Calibri")
        for cell in [ws_sum["A1"], ws_sum["B1"], ws_sum["C1"]]:
            cell.fill = PatternFill("solid", fgColor=PURPLE)

        for i, cat in enumerate(CATEGORIES, 2):
            ws_sum.cell(row=i, column=1, value=cat)
            ws_sum.cell(row=i, column=2, value=f'=COUNTIF(ANNOTATION!E:E,A{i})')
            ws_sum.cell(row=i, column=3, value="auto-calculated")
        ws_sum.cell(row=len(CATEGORIES)+2, column=1, value="TOTAL LABELLED")
        ws_sum.cell(row=len(CATEGORIES)+2, column=2,
                    value=f'=COUNTA(ANNOTATION!E2:E{len(messages)+1})')
        ws_sum.cell(row=len(CATEGORIES)+2, column=1).font = Font(bold=True, name="Calibri")
        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 12
        ws_sum.column_dimensions["C"].width = 22

        out_path = OUT / f"annotation_sheet_A{annotator_num}.xlsx"
        wb.save(out_path)
        print(f"  Saved annotator {annotator_num} sheet → {out_path.name}")

# ── Cohen's Kappa calculator ─────────────────────────────────────────────────

KAPPA_SCRIPT = '''"""
Cohen\'s Kappa Calculator
========================
Run after BOTH annotators complete their sheets:

    python annotation_package/kappa_calculator.py

Reads annotation_sheet_A1.xlsx and annotation_sheet_A2.xlsx,
computes Cohen\'s Kappa, and prints a full agreement report.
"""

from pathlib import Path
import sys

try:
    import openpyxl
    from sklearn.metrics import cohen_kappa_score, confusion_matrix, classification_report
    import pandas as pd
    import numpy as np
except ImportError:
    print("Install dependencies: pip install openpyxl scikit-learn pandas --break-system-packages")
    sys.exit(1)

BASE = Path(__file__).parent

def load_labels(path, col_idx=4):  # col E = index 4 (0-based)
    wb = openpyxl.load_workbook(path)
    ws = wb["ANNOTATION"]
    labels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[col_idx] if len(row) > col_idx else None
        labels.append(str(label).strip() if label else "")
    return labels

print("Loading annotator sheets...")
a1 = load_labels(BASE / "annotation_sheet_A1.xlsx")
a2 = load_labels(BASE / "annotation_sheet_A2.xlsx")

# Filter to only rows where both annotators provided a label
paired = [(l1, l2) for l1, l2 in zip(a1, a2)
          if l1 and l2 and l1 != "None" and l2 != "None"]

if not paired:
    print("No paired labels found. Make sure both annotators have filled in column E.")
    sys.exit(1)

y1 = [p[0] for p in paired]
y2 = [p[1] for p in paired]

kappa = cohen_kappa_score(y1, y2)
agreement = sum(l1 == l2 for l1, l2 in paired) / len(paired) * 100

print(f"\\n{'='*60}")
print(f"INTER-ANNOTATOR AGREEMENT REPORT")
print(f"{'='*60}")
print(f"Messages with both labels:  {len(paired)}")
print(f"Simple agreement:           {agreement:.1f}%")
print(f"Cohen\'s Kappa:              {kappa:.4f}")
print(f"{'='*60}")

if kappa >= 0.80:
    print(f"RESULT: EXCELLENT agreement (kappa >= 0.80) ✓ — ready to publish")
elif kappa >= 0.61:
    print(f"RESULT: SUBSTANTIAL agreement (0.61-0.79) — review disagreements, discuss, re-label")
elif kappa >= 0.41:
    print(f"RESULT: MODERATE agreement (0.41-0.60) — significant discussion and revision needed")
else:
    print(f"RESULT: POOR agreement (<0.41) — guidelines need revision, repeat annotation")

print(f"\\nTarget for publication: kappa >= 0.80")
print(f"{'='*60}")

# Category breakdown
labels_all = sorted(set(y1 + y2))
print(f"\\nPER-CATEGORY BREAKDOWN:")
print(f"{'Category':<28} {'A1':>6} {'A2':>6} {'Agree':>8}")
print("-"*52)
for cat in labels_all:
    c1 = y1.count(cat)
    c2 = y2.count(cat)
    agree = sum(l1 == l2 == cat for l1, l2 in paired)
    print(f"{cat:<28} {c1:>6} {c2:>6} {agree:>8}")

# Disagreement analysis
print(f"\\nDISAGREEMENTS (first 20):")
print(f"{'MSG':<8} {'Annotator 1':<28} {'Annotator 2':<28}")
print("-"*66)
count = 0
for i, (l1, l2) in enumerate(zip(y1, y2)):
    if l1 != l2:
        print(f"MSG{i+1:04d}  {l1:<28} {l2:<28}")
        count += 1
        if count >= 20:
            print("  ... (showing first 20 disagreements only)")
            break

print(f"\\nSaved report. Kappa = {kappa:.4f}")

# Save CSV report
import csv
report_path = BASE / "kappa_report.csv"
with open(report_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["msg_id","annotator1","annotator2","agree"])
    for i, (l1, l2) in enumerate(zip(y1, y2)):
        w.writerow([f"MSG{i+1:04d}", l1, l2, l1==l2])
print(f"Full report saved → {report_path.name}")
'''

# ── Recruitment email ─────────────────────────────────────────────────────────

RECRUITMENT_EMAIL = """SUBJECT: Paid Annotation Task — NLP Research (Bangla, ~4 hrs, $50)

Dear [Name],

I am a Ph.D. student in Computer Science & Engineering at Louisiana State University
working on my dissertation: a measurement study of mobile financial service (MFS) scams
in Bangladesh (bKash, Nagad, Rocket).

I am looking for 2 native Bangla speakers to help label a set of 600 short text messages
as part of the annotation phase of my research.

WHAT YOU WILL DO
  • Read 600 short messages (app reviews, SMS texts)
  • Assign each message one of 7 scam category labels using a dropdown in Excel
  • Work independently at your own pace
  • Estimated time: 3–5 hours (you can split across multiple sessions)

REQUIREMENTS
  • Native Bangla speaker
  • Basic computer skills (Microsoft Excel or Google Sheets)
  • No NLP or security background required — full guidelines provided

COMPENSATION
  • $50 (or equivalent) upon completion — paid via Venmo / Zelle / bKash

WHAT IS PROVIDED
  • Detailed annotation guidelines document
  • Pre-labelled Excel spreadsheet with dropdown menus
  • Full support from me throughout the process

TIMELINE
  • Available to start: [DATE]
  • Deadline: 2 weeks from start

This study has been reviewed by the LSU Institutional Review Board (Protocol No. LSU-IRB-2025-___).
Your name will be acknowledged in the published paper (or kept anonymous, your choice).

If you are interested, please reply to this email or contact me at:
  saidur@lsu.edu

Thank you for your time. I hope you can help!

Best regards,
Md. Saidur Rahman
Ph.D. Student, CSE Division
Louisiana State University
saidur@lsu.edu
"""

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("MFS Scam Annotation — Dataset Preparation")
    print("=" * 50)

    # Load all sources
    print("\nLoading data sources...")
    pool = []
    pool += load_financial_scams()
    pool += load_banglabarta()
    pool += load_app_reviews()
    pool += load_telegram()

    print(f"\nTotal unique messages available: {len(pool)}")

    if len(pool) < TARGET:
        print(f"\n[!] Only {len(pool)} messages available (target: {TARGET}).")
        print("    Will use all available messages.")
        n = len(pool)
    else:
        n = TARGET

    # Sample
    print(f"\nSampling {n} messages (stratified by source)...")
    messages = stratified_sample(pool, n)

    # Source breakdown
    src_counts = Counter(m["source"].split("_")[0] for m in messages)
    print("  Source breakdown:")
    for src, cnt in sorted(src_counts.items()):
        print(f"    {src:<25}: {cnt}")

    # Category breakdown
    cat_counts = Counter(m["prelabel"] for m in messages)
    print("  Pre-label breakdown:")
    for cat, cnt in sorted(cat_counts.items()):
        print(f"    {cat:<30}: {cnt}")

    # Write outputs
    print("\nWriting outputs...")
    write_batch_csv(messages)
    write_excel_sheets(messages)

    # Write kappa calculator
    kappa_path = OUT / "kappa_calculator.py"
    with open(kappa_path, "w", encoding="utf-8") as f:
        f.write(KAPPA_SCRIPT)
    print(f"  Saved Cohen's Kappa calculator → {kappa_path.name}")

    # Write recruitment email
    email_path = OUT / "recruitment_email.txt"
    with open(email_path, "w", encoding="utf-8") as f:
        f.write(RECRUITMENT_EMAIL)
    print(f"  Saved recruitment email → {email_path.name}")

    print("\n" + "="*50)
    print("DONE. Next steps:")
    print("  1. Send recruitment_email.txt to 2 Bangla-speaking colleagues")
    print("  2. Give annotation_sheet_A1.xlsx to Annotator 1")
    print("  3. Give annotation_sheet_A2.xlsx to Annotator 2")
    print("  4. After both finish, run: python annotation_package/kappa_calculator.py")
    print("="*50)

if __name__ == "__main__":
    main()
